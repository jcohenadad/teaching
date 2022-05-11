
# importing openpyxl module
import openpyxl as xl
from loguru import logger


# opening the source excel file
filename = "/Users/julien/Desktop/src.xlsx"
# WARNING!!! Starts at 1
col_id_src = 3
col_val_src = 9
row_start_src = 1
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]

# opening the destination excel file
filename1 = "/Users/julien/Desktop/dest.xlsx"
wb2 = xl.load_workbook(filename1)
ws2 = wb2.active
col_id_dest = 1
col_val_dest = 9
# row_start_dest = 11

# Loop across rows from the src file
n_row_src = ws1.max_row

# copying the cell values from source
# excel file to destination excel file
for i in range(row_start_src+1, n_row_src + 1):
	# Read index value from source file
	id = ws1.cell(row=i, column=col_id_src).value
	# Read value from source file
	val = ws1.cell(row=i, column=col_val_src).value
	# Find index from dest file
	found = False
	for i_row in range(1, ws2.max_row + 1):
		for i_col in range(1, ws2.max_column + 1):
			cell = ws2.cell(i_row, i_col).value
			logger.debug(f"id: {id} | Cell ({i_row}, {i_col}): {cell}")
			if cell == id:
				print("Found!")
				found = True

				break
	if not found:
		logger.error("Not found :-(")

		#
		# # reading cell value from source excel file
		# c = ws1.cell(row = i, column = j)
		#
		# # writing the read value to destination excel file
		# ws2.cell(row = i, column = j).value = c.value

# saving the destination excel file
wb2.save(str(filename1))
