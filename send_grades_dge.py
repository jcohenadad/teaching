import openpyxl, csv, os, re, argparse


def get_parameters():
    parser = argparse.ArgumentParser(description='This script is used for'
        'generating the department reports for GBM')
    parser.add_argument("-d", "--data_xlsx",
                        help="Path to xlsx folder containing the grades",
                        required=True)
    parser.add_argument("-c", "--csv_data",
                        help="Path to csv_file containing the grades",
                        required=True)                    
    args = parser.parse_args()
    return args

def generate_grades_dict(option,path_csv_class):
    option_dict = {
        'PR' : 5,
        'PO' : 6,
        'TP' : 7
    }
    file_csv_class = open(path_csv_class, 'rt')
    reader_csv = csv.reader(file_csv_class)
    buffer_csv_class = []
    for row in reader_csv:
        buffer_csv_class.append(row)
    file_csv_class.close()
    dict_grades_option = {}
    for row_id in range (1,len(buffer_csv_class)):
        student_id = buffer_csv_class[row_id][0]
        grade = buffer_csv_class[row_id][option_dict[option]]
        dict_grades_option[student_id] = grade
    return dict_grades_option


def main(data_xlsx,csv_data):
    """
    Main function
    :param data_xlsx:
    :param csv_data:
    :return:
    """
    data_path = data_xlsx
    path_csv_class = csv_data

    xslx_files = [x for x in os.listdir(data_path) if x.endswith('.xlsx')]

    keys = []
    values = []
    for categ in xslx_files:
        key = " ".join(re.findall("[a-zA-Z]+", categ.split('_')[5]))
        keys.append(key)
        values.append(os.path.join(data_path,categ))

    dict_xslx_file = dict(zip(keys,values))

    for categ in dict_xslx_file.keys():
        wb = openpyxl.load_workbook(dict_xslx_file[categ])
        grades_categ = generate_grades_dict(categ,path_csv_class)
        sheet = wb.get_sheet_by_name('Sheet1')
        ws = wb.worksheets[0]
        dict_dge_option={}
        for i in range(11, 36):
            dict_dge_option[sheet.cell(row=i, column=1).value]=i
        for student in grades_categ.keys():
            if student in dict_dge_option.keys():
                ws.cell(row=dict_dge_option[student], column=4).value = grades_categ[student]
        wb.save(os.path.join(data_path,'Cotes_GBM7904_2020_' + categ + '.xlsx'))

if __name__ == "__main__":
    args = get_parameters()
    main(args.data_xlsx,args.csv_data)