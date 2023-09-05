#!/usr/bin/env python
"""
This script finds the value of a 'source' EXCEL file and insert it in a 'destination' EXCEL file. The matching is done
on a column specified in this script (eg: the 'matricule' of a student). The file should be edited to customize it.
"""

import sys
import openpyxl as xl
from loguru import logger

# Parameters
# Set file names
# GBM8378: https://ele.examen.polymtl.ca/grade/export/ods/index.php?id=1412
# GBM6904: https://moodle.polymtl.ca/grade/export/xls/index.php?id=415
fname_source = "/Users/julien/Dropbox/documents/cours/GBM8378/2023/notes/GBM8378-2023_Final-notes.xlsx"
fname_dest = "Cotes_GBM8378_20231_01_Cours_EF_1_20230205.xlsx"
fname_out = "Cotes_GBM8378_20231_01_Cours_EF_1_20230205_modif.xlsx"
# Set source and destination columns. WARNING!!! Starts at 1
col_id_src = 3  # source column of the students 'matricule'
col_val_src = 7  # source column of the students grade (eg: '5' corresponds to 'E' on the Excel sheet).
row_start_src = 1
col_id_dest = 1  # destination column of the students 'matricule'
col_val_dest = 4  # destination column of the students grade


logger.remove()
logger.add(sys.stderr, level="DEBUG")

# Opening source Excel file
wb1 = xl.load_workbook(fname_source, read_only=True, data_only=True)
ws1 = wb1.active

# opening the destination Excel file
wb2 = xl.load_workbook(fname_dest)
ws2 = wb2.active

# Loop across rows from the src file
# This is a workaround found in https://localcoder.org/openpyxl-max-row-and-max-column-wrongly-reports-a-larger-figure
#  to address the problem of ws1.max_row outputting 'None' sometimes.
for n_row_src, row in enumerate(ws1, 1):
    if all(c.value is None for c in row):
        break


def isvalid_cell(value):
    """
    Check if the cell value fits desired properties
    :param value:
    :return: True if the cell is OK, False otherwise.
    """
    if value is None:
        return False
    # First, check if it is an integer
    if type(value) is int:
        # If it is, check if it has 7 digit
        if len(str(value)) == 7:
            return True
        else:
            return False
    elif type(value) is str:
        # Matricule should be a number
        if not value.isdigit():
            return False
        # Matricule should have 7 digits
        if len(value) == 7:
            return True
        else:
            return False
    else:
        return False


# copying the cell values from source
# Excel file to destination Excel file
for i in range(row_start_src + 1, n_row_src + 1):
    # Read index value from source file
    id = ws1.cell(row=i, column=col_id_src).value
    # Make sure the index is valid
    if not isvalid_cell(id):
        continue
    else:
        # Force it to be a string for subsequent comparison
        id = str(id)
    logger.info(f"id: {id}")
    # Read value from source file
    val = ws1.cell(row=i, column=col_val_src).value
    # Find index from dest file
    found = False
    for i_row in range(1, ws2.max_row + 1):
        for i_col in range(1, ws2.max_column + 1):
            cell = ws2.cell(i_row, i_col).value
            # For it to be a string for the comparison below
            cell = str(cell)
            logger.debug(f"id: {id} | Cell ({i_row}, {i_col}): {cell}")
            if cell == id:
                found = True
                # Assign value in destination cell
                ws2.cell(i_row, col_val_dest).value = val
                logger.info(f"Found matching cell! 🎉")
                break
    if not found:
        logger.error("Not found :-(")

# saving the destination Excel file
wb2.save(fname_out)

logger.info("Job done! 🎉")
