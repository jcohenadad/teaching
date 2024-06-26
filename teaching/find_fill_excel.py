#!/usr/bin/env python
"""
This script finds the value of a 'source' EXCEL file and insert it in a 'destination' EXCEL file. The matching is done
on a column specified in this script (eg: the 'matricule' of a student).

How to use:
- Export files as Excel from Moodle Exam or from Moodle Cours. Make sure the Maximum score is 20.0. To export: go to the 
  page 'Notes', and then click on the combo box on the top left and select 'Exporter'.
- Get the Quotes template from the department. See email "subject: Rapport quotes". Alternatively, download it from GEADE:
  - Clique sur "Cotes et notes"
  - Saisir sigle du cours et trimestre 20233, puis clique sur "Rechercher"
  - Clicque sur "Cotes"
  - Sélectionner le "Contrôle" dans la liste déroulante (devoir, article, etc)
  - Clique sur "Télecharger les cotes"
  - Sélectionner "Exporter cotes" dans "Type de téléchargement"
  - Clique sur "Télécharger"
  - Répéter pour chacun des contrôles
"""

import argparse
import csv
import sys
import openpyxl as xl
from loguru import logger


logger.remove()
logger.add(sys.stderr, level="INFO")

def get_parameters():
    help_description = """
This script finds the value of a 'source' EXCEL file and insert it in a 'destination' EXCEL file. The matching is done
on a column specified in this script (eg: the 'matricule' of a student)."""
    example_commands = """
For GBM8378 (Important: Download CSV with ',' as separator):
find_fill_excel INPUT_CSV OUTPUT_XLS
"""
    parser = argparse.ArgumentParser(description=help_description,
                                     epilog=example_commands,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)

    # Add arguments
    parser.add_argument('file_src', type=str,
                        help='Path to the source EXCEL or CSV file')

    parser.add_argument('file_dest', type=str,
                        help='Name of the destination file')

    parser.add_argument('--col-src-id', type=int, default=3,
                        help='Source column of the student ID (matricule). Starts at 1. Default=1. GBM6904=3')

    parser.add_argument('--col-src-val', type=int, default=9,
                        help='Source column of the student grade, starting at 1. GBM6125=9, GBM6904=7, GBM8378=9')

    parser.add_argument('--row-src-start', type=int, default=0,
                        help='Starting row in source file, starting at 1')

    parser.add_argument('--col-dest-id', type=int, default=1,
                        help='Destination column of the student ID (matricule), starting at 1')

    parser.add_argument('--col-dest-val', type=int, default=4,
                        help='Destination column of the student grade, starting at 1. GBM6904=4')
    parser.add_argument('--debug', action='store_true', help='Debug mode')

    args = parser.parse_args()
    return args


def isvalid_cell(value):
    """
    Check if the cell value fits desired properties
    :param value:
    :return: True if the cell is OK, False otherwise.
    """
    if value is None:
        return False
    # First, check if it is an integer
    if type(value) is int:
        # If it is, check if it has 7 digit
        if len(str(value)) == 7:
            return True
        else:
            return False
    elif type(value) is str:
        # Matricule should be a number
        if not value.isdigit():
            return False
        # Matricule should have 7 digits
        if len(value) == 7:
            return True
        else:
            return False
    else:
        return False


def main():
    """
    Copy the cell values from source file to destination Excel file
    """

    # get input parameters
    args = get_parameters()
    fname_source = args.file_src
    fname_dest = args.file_dest
    col_src_id = args.col_src_id
    col_src_val = args.col_src_val
    row_src_start = args.row_src_start
    col_dest_id = args.col_dest_id
    col_dest_val = args.col_dest_val
    logger_mode = args.debug

    # Set logger mode
    if logger_mode:
        logger.remove()
        logger.add(sys.stderr, level="DEBUG")

    # Figure out if the source file is an Excel or CSV file
    if fname_source.endswith(('.xlsx', '.xls')):
        logger.info("Source file is an Excel file")
        # Opening source Excel file
        wb1 = xl.load_workbook(fname_source, read_only=True, data_only=True)
        ws1 = wb1.active
    elif fname_source.endswith('.csv'):
        logger.info("Source file is a CSV file")
        # Opening source CSV file
        wb1 = xl.Workbook()
        ws1 = wb1.active
        with open(fname_source, 'r') as f:
            for row in csv.reader(f, quotechar='"', delimiter=',', quoting=csv.QUOTE_ALL, skipinitialspace=True):
                ws1.append(row)
    else:
        logger.error("Source file should be an Excel or CSV file")
        sys.exit(1)

    # opening the destination Excel file
    wb2 = xl.load_workbook(fname_dest)
    ws2 = wb2.active

    # Loop across rows from the src file
    # This is a workaround found in https://localcoder.org/openpyxl-max-row-and-max-column-wrongly-reports-a-larger-figure
    #  to address the problem of ws1.max_row outputting 'None' sometimes.
    for n_row_src, row in enumerate(ws1, 1):
        if all(c.value is None for c in row):
            break

    for i in range(row_src_start + 1, n_row_src + 1):
        # Read index value from source file
        id = ws1.cell(row=i, column=col_src_id).value
        logger.debug(f"Source: i={i}, id={id}")
        # Make sure the index is valid
        if not isvalid_cell(id):
            continue
        else:
            # Force it to be a string for subsequent comparison
            id = str(id)
        # Read value from source file
        val = ws1.cell(row=i, column=col_src_val).value
        # Replace ',' by '.' in the value
        if type(val) is str:
            val = val.replace(',', '.')
        # Make sure the value is valid and convert it to float. If not, skip it.
        if val is None:
            logger.info(f"❌ Source: matricule={id}, value={val} (ignoring)")
            continue
        try:
            val = float(val)
        except ValueError:
            logger.info(f"❌ Source: matricule={id}, value={val} (ignoring)")
            continue
        logger.info(f"✅ Source: matricule={id}, value={val}")
        # Find index from dest file
        found = False
        for i_row in range(1, ws2.max_row + 1):
            # Fetch matricule on destination file
            cell = str(ws2.cell(i_row, col_dest_id).value)
            logger.debug(f"Destination: i_row={i_row}, cell={cell}")
            if cell == id:
                found = True
                # Assign value in destination cell
                ws2.cell(i_row, col_dest_val).value = val
                logger.info(f"Found matching cell! 🎉")
                break
        if not found:
            logger.error("Not found :-(")

    # saving the destination Excel file
    wb2.save(fname_dest)

    logger.info("Job done! 🎉")

if __name__ == '__main__':
    main()
